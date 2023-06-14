import os
import ast
import math
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


from doctors_result import doctor_result_dict

#вместо этой функции необходимо написать работающий код
def ctg_analyze(x_coords, y_coords):
    x_cor = x_coords
    y_cor = y_coords
    def clear_coords(x_coords,y_coords):
            dictioanry = dict(zip(x_coords, y_coords))
            for k,v in list(dictioanry.items()):
                 if v == 0 or v >180:
                    dictioanry.pop(k)
            return dictioanry
    diction = clear_coords(x_coords,y_coords)
    x_coords = list(diction.keys())
    y_coords = list(diction.values())
    avg_bas = []
    count_list = []       
    difr = 50
    for i in range(len(y_coords)):
        if i % difr==0 and i!=0:
            sr=sum(count_list) / len(count_list)
            avg_bas.append(sr)
            count_list.clear()
        else:
            count_list.append(y_coords[i])         
    avg_bas_per=sum(avg_bas) / len(avg_bas)
    y_coords=[0 if i>avg_bas_per+10 or i<avg_bas_per-10 else i for i in y_coords]
    diction = clear_coords(x_coords,y_coords)
    x_coords = list(diction.keys())
    y_coords = list(diction.values())
    y_avg=sum(y_coords) / len(y_coords)
    dif = []
    difr = 200
    big_list = []
    small_list = []
    avg_list = []
    avg = []
    for i in range(len(y_coords)-200):
        big = 0
        small = 1000
        if i % difr == 0 and i != 0:
            for j in range(i, i+difr-1):
                avg.append(y_coords[j])
                if y_coords[j]>big:
                    big = y_coords[j]
                if y_coords[j]<small:
                    small = y_coords[j]
            a=big-small
            big_list.append(big)
            small_list.append(small)
            avg_list.append(sum(avg) / len(avg))
            avg.clear()
            dif.append(abs(a))
    amp = []  
    avg_dif=sum(dif) / len(dif)
    for i in range(len(avg_list)):
        if avg_list[i] > avg_dif:
            amp.append(abs(big_list[i]-y_avg)) 
        else:
            amp.append(y_avg-small_list[i])
    x_cor = list(range(1,len(dif)+1))
    up_list = []
    down_list = []
    for i in range(len(amp)):
        if amp[i] > 15:
            up_list.append(i)
        else:
            down_list.append(i)
    y_coords_true = []
    count = down_list[0]
    count2 = 0
    for i in range(len(y_coords)):
        if i % difr == 0 and i != 0:
            if count2 == len(down_list ):
                break
            if count == down_list[count2]:
                for j in range(i, i+difr-1):
                    y_coords_true.append(y_coords[j])
                count2 = count2+1
            count = count+1
    #
    ist_base = sum(y_coords_true) / len(y_coords_true)
    #
    count = down_list[0]
    count2 = 0
    amplit = []
    for i in range(len(down_list)):
        amplit.append(big_list[i]-small_list[i])
    #
    amplit =  sum(amplit) / len(amplit)
    #
    if amplit > 29.6 or amplit < 9.6:
        return "плохое"
    else:
        if ist_base < 126.7:
            return "плохое"
        else:
            return "хорошее"


if __name__ == '__main__':

    directory = 'ctg_files'
    program_result_dict = {}
    start_time = datetime.now()
    #проходим циклом по предоставленным файлам с массивами данных по КТГ
    filename_list = os.listdir(directory)
    filename_list.sort(key=lambda x: int(x[:-4]))
    for filename in filename_list:
        f = os.path.join(directory, filename)
        if os.path.isfile(f):
            file = open(f, 'r')
            graph_list = ast.literal_eval(file.read())
            #преобразуем данные в pandas dataframe для дальнейшей обработки
            #преобразовывать в dataframe необязательно, если имеются другие решения можете реализовать их
            x_coords = [i.get('Key') for i in graph_list]
            y_coords = [i.get('Value') for i in graph_list]
            df_all_coords = pd.DataFrame.from_dict({'x': x_coords, 
                                                    'y': y_coords, 
                                                    })
            x_list = x_coords
            y_list = y_coords
            '''plt.bar(x_list, y_list)
            plt.show()'''
            #//////////////////////////////////////////////////////////////////////////////////////////////////////////////
            '''здесь вызывается исполнение функции оценивающей КТГ
            при написании кода рекомендуется использование matplotlib или аналоги для визуализации графика, это поможет
            писать весь код в одной функции необязательно - хорошая читаемость кода приветствуется
            программа в результате должна вернуть строку 'хорошее' или 'плохое'
            '''
            program_result = ctg_analyze(x_coords, y_coords)
            #//////////////////////////////////////////////////////////////////////////////////////////////////////////////


            #записывается результирующий словарь с ключами идентичными словарю doctor_result_dict для дальнейшего сравнения
            program_result_dict[filename] = program_result

    #считаем среднее время выполнения оценки одного КТГ
    average_time = (datetime.now() - start_time) / len(os.listdir(directory))
    print(f'среднее время выполнения оценки одного КТГ - {average_time}')

    #считаем количество совпадений программы с врачом
    number_of_matches = 0
    for res in program_result_dict:
        if program_result_dict[res] == doctor_result_dict[res]:
            number_of_matches += 1
    print(f'совпадений программы с врачом {number_of_matches} из 100')

    #в таблицу сохраняется результат
    #в ней можно будет более подробно рассмотреть общую картину того в каких случаях расхождения между врачом и программой
    wb = load_workbook('ctg.xlsx')
    del wb['Sheet1']
    ws = wb.create_sheet('Sheet1')
    for res in program_result_dict:
        ws.append([res, doctor_result_dict[res], program_result_dict[res]])
    wb.save('ctg.xlsx')